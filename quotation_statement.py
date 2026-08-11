"""Quotation and transaction-statement calculation and XLSX generation.

The module deliberately has no Qt dependency.  It owns the fixed pricing
rules, per-unit VAT rounding, and the small amount of template manipulation
needed by both documents.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import os
import re
import subprocess

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter, range_boundaries


WON = Decimal("1")
VAT_DIVISOR = Decimal("1.1")
SHIPPING_GROSS = 3000
SHIPPING_SUPPLY = 2727
SHIPPING_TAX = 273
NEGOTIATION_LIMIT = 5_000_000
WON_NUMBER_FORMAT = '"₩"#,##0'
PAYMENT_METHODS = {
    "직거래": "직접 거래",
    "네이버": "네이버 스토어 거래",
    "쿠팡": "쿠팡 스토어 거래",
    "G마켓": "G마켓 거래",
    "옥션": "옥션 거래",
    "11번가": "11번가 거래",
}


class DocumentValidationError(ValueError):
    """Raised when document input is incomplete or invalid."""


class NegotiationRequired(DocumentValidationError):
    """Raised when an order over five million won is not confirmed."""


class PdfExportError(RuntimeError):
    """Raised when Excel cannot export the generated workbook as PDF."""


def _write_pdf_export_log(source: Path, target: Path, status: str, detail: str = "") -> None:
    """Append diagnostics without making a document export fail because logging is unavailable."""
    try:
        log_path = source.parent / "pdf_export.log"
        lines = [
            f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {status}",
            f"source={source.name}",
            f"target={target.name}",
        ]
        if detail:
            lines.append(detail.strip())
        with log_path.open("a", encoding="utf-8") as stream:
            stream.write(" | ".join(lines) + "\n")
    except OSError:
        pass


@dataclass(frozen=True)
class ItemInput:
    document_name: str
    specification: str = ""
    gross_unit_price: str | int | Decimal = 0
    quantity: int = 1
    api_name: str = ""


@dataclass(frozen=True)
class CalculatedItem:
    document_name: str
    specification: str
    quantity: int
    original_gross_unit_price: int
    gross_unit_price: int
    supply_unit_price: int
    tax_unit_price: int
    supply_amount: int
    tax_amount: int

    @property
    def gross_amount(self) -> int:
        return self.supply_amount + self.tax_amount


@dataclass(frozen=True)
class CalculationResult:
    items: tuple[CalculatedItem, ...]
    goods_total_before_discount: int
    discount_rate: Decimal
    discount_amount: int
    shipping_gross: int
    supply_total: int
    tax_total: int
    grand_total: int
    negotiated: bool


def round_won(value: Decimal) -> int:
    return int(value.quantize(WON, rounding=ROUND_HALF_UP))


def _positive_won(value: str | int | Decimal, label: str) -> int:
    try:
        cleaned = str(value).replace(",", "").strip()
        amount = round_won(Decimal(cleaned))
    except (InvalidOperation, ValueError):
        raise DocumentValidationError(f"{label}은(는) 올바른 금액이어야 합니다.") from None
    if amount < 1:
        raise DocumentValidationError(f"{label}은(는) 1원 이상이어야 합니다.")
    return amount


def discount_rate_for(goods_total: int) -> Decimal:
    if goods_total <= 100_000:
        return Decimal("0")
    if goods_total <= 2_000_000:
        return Decimal("0.05")
    if goods_total <= NEGOTIATION_LIMIT:
        return Decimal("0.10")
    raise NegotiationRequired("500만 원 초과 주문은 별도 협의가 필요합니다.")


def calculate_document(
    items: list[ItemInput] | tuple[ItemInput, ...], *, negotiated: bool = False
) -> CalculationResult:
    if not items:
        raise DocumentValidationError("품목을 한 개 이상 입력해주세요.")

    normalized: list[tuple[ItemInput, int, int]] = []
    goods_total = 0
    for index, item in enumerate(items, start=1):
        name = str(item.document_name or "").strip()
        if not name:
            raise DocumentValidationError(f"{index}번 품목의 문서용 상품명을 입력해주세요.")
        try:
            quantity = int(item.quantity)
        except (TypeError, ValueError):
            raise DocumentValidationError(f"{index}번 품목의 수량은 정수여야 합니다.") from None
        if quantity < 1 or str(item.quantity).strip() != str(quantity):
            raise DocumentValidationError(f"{index}번 품목의 수량은 1개 이상 정수여야 합니다.")
        gross_unit = _positive_won(item.gross_unit_price, f"{index}번 품목 단가")
        normalized.append((item, gross_unit, quantity))
        goods_total += gross_unit * quantity

    if negotiated:
        rate = Decimal("0")
    elif goods_total > NEGOTIATION_LIMIT:
        raise NegotiationRequired("500만 원 초과 주문은 협의 완료 확인 후 생성할 수 있습니다.")
    else:
        rate = discount_rate_for(goods_total)

    calculated: list[CalculatedItem] = []
    supply_total = 0
    tax_total = 0
    discounted_goods_total = 0
    multiplier = Decimal("1") - rate
    for item, original_gross, quantity in normalized:
        gross_unit = round_won(Decimal(original_gross) * multiplier)
        supply_unit = round_won(Decimal(gross_unit) / VAT_DIVISOR)
        tax_unit = gross_unit - supply_unit
        supply_amount = supply_unit * quantity
        tax_amount = tax_unit * quantity
        calculated.append(
            CalculatedItem(
                document_name=str(item.document_name).strip(),
                specification=str(item.specification or "").strip(),
                quantity=quantity,
                original_gross_unit_price=original_gross,
                gross_unit_price=gross_unit,
                supply_unit_price=supply_unit,
                tax_unit_price=tax_unit,
                supply_amount=supply_amount,
                tax_amount=tax_amount,
            )
        )
        supply_total += supply_amount
        tax_total += tax_amount
        discounted_goods_total += gross_unit * quantity

    shipping = SHIPPING_GROSS if goods_total <= 100_000 and not negotiated else 0
    if shipping:
        supply_total += SHIPPING_SUPPLY
        tax_total += SHIPPING_TAX

    grand_total = supply_total + tax_total
    return CalculationResult(
        items=tuple(calculated),
        goods_total_before_discount=goods_total,
        discount_rate=rate,
        discount_amount=goods_total - discounted_goods_total,
        shipping_gross=shipping,
        supply_total=supply_total,
        tax_total=tax_total,
        grand_total=grand_total,
        negotiated=negotiated,
    )


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _insert_item_rows(ws, insert_at: int, count: int, source_row: int, merge_columns) -> None:
    if count <= 0:
        return

    merged = [str(cell_range) for cell_range in ws.merged_cells.ranges]
    for cell_range in merged:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        if max_row >= insert_at:
            ws.unmerge_cells(cell_range)

    dimensions = {
        row: copy(dimension)
        for row, dimension in ws.row_dimensions.items()
        if row >= insert_at
    }
    for row in dimensions:
        del ws.row_dimensions[row]

    ws.insert_rows(insert_at, count)

    for row, dimension in dimensions.items():
        dimension.index = row + count
        ws.row_dimensions[row + count] = dimension

    for cell_range in merged:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        if min_row >= insert_at:
            min_row += count
            max_row += count
        elif max_row >= insert_at:
            max_row += count
        ws.merge_cells(
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

    for row in range(insert_at, insert_at + count):
        _copy_row_style(ws, source_row, row)
        for start_col, end_col in merge_columns:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)


def _safe_component(value: str) -> str:
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", str(value or "").strip())
    text = re.sub(r"\s+", "_", text).strip(" ._")
    return text or "미입력"


def _output_path(output_dir: Path, kind: str, organization: str, name: str, trade_date: date) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = f"{kind}_{_safe_component(organization)}_{_safe_component(name)}_{trade_date:%Y.%m.%d}"
    path = output_dir / f"{stem}.xlsx"
    index = 1
    while path.exists() or path.with_suffix(".pdf").exists():
        path = output_dir / f"{stem}_({index}).xlsx"
        index += 1
    return path


def _validate_header(organization: str, name: str, trade_date: date) -> None:
    if not str(organization or "").strip():
        raise DocumentValidationError("소속명을 입력해주세요.")
    if not str(name or "").strip():
        raise DocumentValidationError("성명을 입력해주세요.")
    if not isinstance(trade_date, date):
        raise DocumentValidationError("거래일자를 확인해주세요.")


def _write_lines(ws, start_row: int, result: CalculationResult, columns: dict[str, str]) -> int:
    row = start_row
    for item in result.items:
        ws[f"{columns['name']}{row}"] = item.document_name
        if len(item.document_name) > 28:
            name_cell = ws[f"{columns['name']}{row}"]
            alignment = copy(name_cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "center"
            name_cell.alignment = alignment
            lines = min(3, (len(item.document_name) + 27) // 28)
            ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 18, 16 * lines)
        ws[f"{columns['spec']}{row}"] = item.specification
        ws[f"{columns['quantity']}{row}"] = item.quantity
        ws[f"{columns['unit']}{row}"] = item.supply_unit_price
        ws[f"{columns['supply']}{row}"] = item.supply_amount
        ws[f"{columns['tax']}{row}"] = item.tax_amount
        row += 1
    if result.shipping_gross:
        ws[f"{columns['name']}{row}"] = "배송비"
        ws[f"{columns['quantity']}{row}"] = 1
        ws[f"{columns['unit']}{row}"] = SHIPPING_SUPPLY
        ws[f"{columns['supply']}{row}"] = SHIPPING_SUPPLY
        ws[f"{columns['tax']}{row}"] = SHIPPING_TAX
        row += 1
    return row


def _prepare_sheet(template: Path, line_count: int, *, quote: bool):
    if not template.exists():
        raise FileNotFoundError(f"문서 템플릿을 찾을 수 없습니다: {template}")
    wb = load_workbook(template)
    ws = wb.active
    if quote:
        capacity, total_row, source_row = 7, 23, 22
        merge_columns = ((2, 6), (8, 9), (11, 12), (14, 16), (18, 20))
    else:
        capacity, total_row, source_row = 6, 16, 15
        merge_columns = ((2, 6), (8, 9), (11, 12), (13, 15), (17, 19), (20, 23))
    extra = max(0, line_count - capacity)
    _insert_item_rows(ws, total_row, extra, source_row, merge_columns)
    return wb, ws, total_row + extra, extra


def _write_quote_amount_label(ws) -> None:
    """원본 견적서처럼 제목과 괄호 설명의 글자 크기를 구분한다."""
    base_font = ws["B12"].font
    ws["B12"] = CellRichText(
        TextBlock(
            InlineFont(rFont=base_font.name, sz=14, b=base_font.b, i=base_font.i),
            "견적금액 ",
        ),
        TextBlock(
            InlineFont(rFont=base_font.name, sz=9, b=base_font.b, i=base_font.i),
            "(공급가액 + 세액) ",
        ),
    )


def _write_statement_total_label(ws) -> None:
    """원본 거래명세서처럼 합계와 부가세 설명의 글자 크기를 구분한다."""
    base_font = ws["B7"].font
    ws["B7"] = CellRichText(
        TextBlock(
            InlineFont(rFont=base_font.name, sz=14, b=base_font.b, i=base_font.i),
            "합계 ",
        ),
        TextBlock(
            InlineFont(rFont=base_font.name, sz=9, b=base_font.b, i=base_font.i),
            "(부가세 포함) ",
        ),
    )


def generate_document(
    document_type: str,
    organization: str,
    name: str,
    trade_date: date,
    items: list[ItemInput] | tuple[ItemInput, ...],
    *,
    negotiated: bool = False,
    payment_method: str = "직거래",
    delivery_term: str = "",
    templates_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
) -> tuple[Path, CalculationResult]:
    _validate_header(organization, name, trade_date)
    result = calculate_document(items, negotiated=negotiated)
    base = Path(__file__).resolve().parent
    templates = Path(templates_dir) if templates_dir else base / "templates"
    output = Path(output_dir) if output_dir else base / "output"
    is_quote = document_type == "견적서"
    if not is_quote and document_type != "거래명세서":
        raise DocumentValidationError("문서 종류를 확인해주세요.")

    line_count = len(result.items) + (1 if result.shipping_gross else 0)
    template_name = "quotation_template.xlsx" if is_quote else "transaction_statement_template.xlsx"
    wb, ws, total_row, extra = _prepare_sheet(templates / template_name, line_count, quote=is_quote)
    try:
        if is_quote:
            try:
                payment_text = PAYMENT_METHODS[payment_method]
            except KeyError:
                raise DocumentValidationError("결제 방법을 확인해주세요.") from None
            ws["B4"] = str(organization).strip()
            ws["D6"] = f"{str(name).strip()}님 귀하"
            ws["D7"] = trade_date.strftime("%Y.%m.%d")
            _write_quote_amount_label(ws)
            ws["R12"] = result.grand_total
            _write_lines(
                ws,
                16,
                result,
                {"name": "B", "spec": "H", "quantity": "K", "unit": "N", "supply": "R", "tax": "V"},
            )
            ws[f"R{total_row}"] = result.supply_total
            ws[f"V{total_row}"] = result.tax_total
            for cell in (ws["R12"], ws[f"R{total_row}"], ws[f"V{total_row}"]):
                cell.number_format = WON_NUMBER_FORMAT
            ws[f"B{total_row + 3}"] = f"기본 납기 : {str(delivery_term).strip() or '결제 후 즉시'}"
            payment_row = total_row + 4
            ws[f"B{payment_row}"] = f"결제 방법 : {payment_text}"
            # 할인 항목을 제거한 빈 행도 숨겨 Excel/PDF에 남지 않게 한다.
            blank_row = payment_row + 1
            ws.unmerge_cells(f"B{blank_row}:V{blank_row}")
            ws.row_dimensions[blank_row].hidden = True
            ws.print_area = f"A1:V{payment_row}"
            # Keep quotation content at the top of the A4 page when few items exist.
            ws.print_options.verticalCentered = False
        else:
            ws["B5"] = f"{str(organization).strip()} {str(name).strip()}님 귀하"
            # 날짜 일련번호를 표시하는 일부 뷰어도 있으므로 문서에는 고정 문자열로 기록한다.
            ws["B6"] = trade_date.strftime("%Y.%m.%d")
            _write_statement_total_label(ws)
            ws["H7"] = result.grand_total
            _write_lines(
                ws,
                10,
                result,
                {"name": "B", "spec": "H", "quantity": "K", "unit": "M", "supply": "Q", "tax": "T"},
            )
            ws[f"Q{total_row}"] = result.supply_total
            ws[f"T{total_row}"] = result.tax_total
            ws.print_area = f"A1:W{18 + extra}"

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0 if extra else 1
        path = _output_path(output, document_type, organization, name, trade_date)
        wb.save(path)
    finally:
        wb.close()
    return path, result


def export_xlsx_to_pdf(xlsx_path: str | Path) -> Path:
    """A4 기준 Excel 인쇄영역을 PDF로 내보내고 기본 프린터를 복원한다."""
    source = Path(xlsx_path).resolve()
    if not source.is_file():
        raise PdfExportError(f"엑셀 파일을 찾을 수 없습니다: {source}")
    target = source.with_suffix(".pdf")
    script = r'''
$ErrorActionPreference = 'Stop'
$originalPrinter = $null
$pdfPrinter = $null
$excel = $null
$book = $null
function Invoke-ComCleanupWithRetry {
    param([string]$Name, [scriptblock]$Action)
    foreach ($attempt in 1..3) {
        try {
            & $Action
            Write-Output "cleanup=$Name attempt=$attempt result=success"
            return
        } catch {
            $hresult = $_.Exception.HResult
            Write-Output "cleanup=$Name attempt=$attempt hresult=$hresult error=$($_.Exception.Message)"
            if ($attempt -lt 3 -and $hresult -eq -2147418111) {
                Start-Sleep -Milliseconds (250 * $attempt)
                continue
            }
            return
        }
    }
}
try {
    $pdfPrinter = Get-CimInstance Win32_Printer -Filter "Name = 'Microsoft Print to PDF'"
    if ($null -eq $pdfPrinter) {
        throw 'A4 PDF 출력을 위한 Microsoft Print to PDF 프린터를 찾을 수 없습니다.'
    }
    $originalPrinter = Get-CimInstance Win32_Printer | Where-Object Default | Select-Object -First 1
    if ($null -eq $originalPrinter) {
        throw '현재 Windows 기본 프린터를 확인할 수 없습니다.'
    }
    & "$env:SystemRoot\System32\rundll32.exe" 'printui.dll,PrintUIEntry' '/y' '/n' $pdfPrinter.Name
    Start-Sleep -Milliseconds 500
    $selectedPrinter = Get-CimInstance Win32_Printer | Where-Object Default | Select-Object -First 1
    if ($null -eq $selectedPrinter -or $selectedPrinter.Name -ne $pdfPrinter.Name) {
        throw 'A4 PDF 출력용 프린터를 선택하지 못했습니다.'
    }
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $book = $excel.Workbooks.Open($env:EASY_FULFILL_XLSX_PATH, 0, $true)
    $sheet = $book.Worksheets.Item(1)
    $sheet.PageSetup.PaperSize = 9
    $sheet.PageSetup.Orientation = 1
    $sheet.PageSetup.Zoom = $false
    $sheet.PageSetup.FitToPagesWide = 1
    $sheet.PageSetup.FitToPagesTall = 1
    $sheet.PageSetup.CenterVertically = $false
    $book.ExportAsFixedFormat(0, $env:EASY_FULFILL_PDF_PATH, 0, $true, $false)
} finally {
    if ($book -ne $null) {
        Invoke-ComCleanupWithRetry 'Workbook.Close' { $book.Close($false) }
        try { [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($book) }
        catch { Write-Output "cleanup=Workbook.Release error=$($_.Exception.Message)" }
    }
    if ($excel -ne $null) {
        Invoke-ComCleanupWithRetry 'Excel.Quit' { $excel.Quit() }
        try { [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($excel) }
        catch { Write-Output "cleanup=Excel.Release error=$($_.Exception.Message)" }
    }
    if ($originalPrinter -ne $null) {
        & "$env:SystemRoot\System32\rundll32.exe" 'printui.dll,PrintUIEntry' '/y' '/n' $originalPrinter.Name
        Start-Sleep -Milliseconds 500
    }
}
'''
    env = os.environ.copy()
    env["EASY_FULFILL_XLSX_PATH"] = str(source)
    env["EASY_FULFILL_PDF_PATH"] = str(target)
    try:
        completed = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", script],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=60,
            env=env,
        )
    except FileNotFoundError as exc:
        _write_pdf_export_log(source, target, "failed", "PowerShell was not found.")
        raise PdfExportError("Windows PowerShell을 찾을 수 없습니다.") from exc
    except subprocess.TimeoutExpired as exc:
        _write_pdf_export_log(source, target, "failed", "Excel PDF conversion timed out after 60 seconds.")
        raise PdfExportError("Excel PDF 변환 시간이 초과되었습니다.") from exc
    detail = "\n".join(part for part in (completed.stdout, completed.stderr) if part).strip()
    if completed.returncode != 0:
        if target.is_file() and target.stat().st_size > 0:
            _write_pdf_export_log(
                source,
                target,
                "completed_with_cleanup_warning",
                f"returncode={completed.returncode}\n{detail}",
            )
            return target
        _write_pdf_export_log(source, target, "failed", f"returncode={completed.returncode}\n{detail}")
        detail = detail or "알 수 없는 오류"
        raise PdfExportError(f"Excel PDF 변환 실패: {detail}")
    if not target.is_file() or target.stat().st_size == 0:
        _write_pdf_export_log(source, target, "failed", "Excel finished without creating a PDF file.")
        raise PdfExportError("Excel이 PDF 파일을 만들지 못했습니다.")
    _write_pdf_export_log(source, target, "completed", detail)
    return target
