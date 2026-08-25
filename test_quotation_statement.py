import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path
import subprocess
from unittest.mock import patch

from openpyxl import load_workbook

from quotation_statement import (
    DocumentValidationError,
    ItemInput,
    NegotiationRequired,
    PAYMENT_METHODS,
    WON_NUMBER_FORMAT,
    calculate_document,
    export_xlsx_to_pdf,
    generate_document,
)
import naver_commerce
import coupang_commerce


ROOT = Path(__file__).resolve().parent
TEMPLATES = ROOT / "templates"


def item(price, quantity=1, name="테스트 상품", spec="규격"):
    return ItemInput(name, spec, str(price), quantity)


class CalculationTests(unittest.TestCase):
    def test_payment_method_order_and_labels(self):
        self.assertEqual(list(PAYMENT_METHODS.items()), [
            ("직거래", "직접 거래"),
            ("네이버", "네이버 스토어 거래"),
            ("쿠팡", "쿠팡 스토어 거래"),
            ("G마켓", "G마켓 거래"),
            ("옥션", "옥션 거래"),
            ("11번가", "11번가 거래"),
        ])

    def test_discount_boundaries(self):
        cases = [
            (100_000, Decimal("0"), 3_000),
            (100_001, Decimal("0.05"), 0),
            (2_000_000, Decimal("0.05"), 0),
            (2_000_001, Decimal("0.10"), 0),
            (5_000_000, Decimal("0.10"), 0),
        ]
        for price, rate, shipping in cases:
            with self.subTest(price=price):
                result = calculate_document([item(price)])
                self.assertEqual(result.discount_rate, rate)
                self.assertEqual(result.shipping_gross, shipping)

    def test_over_five_million_requires_confirmation(self):
        with self.assertRaises(NegotiationRequired):
            calculate_document([item(5_000_001)])
        result = calculate_document([item(5_000_001)], negotiated=True)
        self.assertTrue(result.negotiated)
        self.assertEqual(result.discount_rate, Decimal("0"))
        self.assertEqual(result.grand_total, 5_000_001)

        # 협의 후 입력한 확정 단가가 500만 원 아래로 내려가도 추가 할인을 하지 않는다.
        result = calculate_document([item(4_900_000)], negotiated=True)
        self.assertEqual(result.discount_rate, Decimal("0"))
        self.assertEqual(result.grand_total, 4_900_000)

    def test_four_hundred_unit_example(self):
        result = calculate_document([item(10_000, 400)])
        calculated = result.items[0]
        self.assertEqual(result.discount_rate, Decimal("0.10"))
        self.assertEqual(calculated.gross_unit_price, 9_000)
        self.assertEqual(calculated.supply_unit_price, 8_182)
        self.assertEqual(calculated.tax_unit_price, 818)
        self.assertEqual(calculated.supply_amount, 3_272_800)
        self.assertEqual(calculated.tax_amount, 327_200)

    def test_shipping_vat_split(self):
        result = calculate_document([item(9_000)])
        self.assertEqual(result.shipping_gross, 3_000)
        self.assertEqual(result.supply_total, 8_182 + 2_727)
        self.assertEqual(result.tax_total, 818 + 273)

    def test_free_shipping_removes_low_order_shipping(self):
        result = calculate_document([item(9_000)], free_shipping=True)
        self.assertEqual(result.shipping_gross, 0)
        self.assertEqual(result.supply_total, 8_182)
        self.assertEqual(result.tax_total, 818)
        self.assertEqual(result.grand_total, 9_000)

    def test_manual_discount_rate_overrides_automatic_rate(self):
        result = calculate_document(
            [item(30_000, 4)], discount_rate_override=Decimal("0.07")
        )
        self.assertEqual(result.discount_rate, Decimal("0.07"))
        self.assertEqual(result.discount_amount, 8_400)
        self.assertEqual(result.grand_total, 111_600)

    def test_zero_discount_override_skips_automatic_discount(self):
        result = calculate_document(
            [item(30_000, 4)], discount_rate_override=Decimal("0")
        )
        self.assertEqual(result.discount_rate, Decimal("0"))
        self.assertEqual(result.discount_amount, 0)
        self.assertEqual(result.grand_total, 120_000)

    def test_manual_discount_rate_must_be_a_valid_percentage(self):
        for invalid_rate in (
            Decimal("-0.01"), Decimal("1.01"), Decimal("NaN"), Decimal("0.075")
        ):
            with self.subTest(rate=invalid_rate), self.assertRaises(DocumentValidationError):
                calculate_document([item(30_000)], discount_rate_override=invalid_rate)

    def test_over_limit_manual_discount_still_requires_confirmation(self):
        with self.assertRaises(NegotiationRequired):
            calculate_document(
                [item(5_000_001)], discount_rate_override=Decimal("0.12")
            )
        result = calculate_document(
            [item(5_000_001)],
            negotiated=True,
            discount_rate_override=Decimal("0.12"),
        )
        self.assertTrue(result.negotiated)
        self.assertEqual(result.discount_rate, Decimal("0.12"))

    def test_one_rate_is_applied_to_all_items(self):
        result = calculate_document([item(70_000, name="A"), item(50_001, name="B")])
        self.assertEqual(result.discount_rate, Decimal("0.05"))
        self.assertEqual([x.gross_unit_price for x in result.items], [66_500, 47_501])

    def test_naver_order_import_keeps_actual_total_without_extra_discount_or_shipping(self):
        result = calculate_document([
            ItemInput("주문 상품", gross_unit_price=5_000, quantity=2, gross_amount_override=9_999)
        ], order_import=True, discount_rate_override=Decimal("0.20"))
        self.assertEqual(result.discount_rate, Decimal("0"))
        self.assertEqual(result.discount_amount, 0)
        self.assertEqual(result.shipping_gross, 0)
        self.assertEqual(result.grand_total, 9_999)
        self.assertEqual(result.items[0].supply_amount + result.items[0].tax_amount, 9_999)

    def test_naver_order_import_does_not_require_over_limit_negotiation(self):
        result = calculate_document([
            ItemInput("주문 상품", gross_unit_price=6_000_000, quantity=1, gross_amount_override=6_000_000)
        ], order_import=True)
        self.assertEqual(result.grand_total, 6_000_000)
        self.assertEqual(result.discount_rate, Decimal("0"))

    def test_naver_order_import_adds_only_the_actual_customer_shipping_fee(self):
        result = calculate_document([
            ItemInput("주문 상품", gross_unit_price=5_000, quantity=2, gross_amount_override=9_999)
        ], order_import=True, shipping_gross_override=2_500)
        self.assertEqual(result.shipping_gross, 2_500)
        self.assertEqual(result.shipping_supply, 2_273)
        self.assertEqual(result.shipping_tax, 227)
        self.assertEqual(result.grand_total, 12_499)


class WorkbookTests(unittest.TestCase):
    def test_templates_do_not_contain_sample_customer_or_items(self):
        forbidden = ("글앤정", "임태규", "김형래", "ESP32  DEVKIT", "LD2410C", "Super Mini")
        for template in TEMPLATES.glob("*.xlsx"):
            wb = load_workbook(template, data_only=False)
            try:
                text = "\n".join(
                    str(cell.value)
                    for ws in wb.worksheets
                    for row in ws.iter_rows()
                    for cell in row
                    if cell.value not in (None, "")
                )
                for value in forbidden:
                    self.assertNotIn(value, text, f"{template.name}에 샘플값 잔존")
                self.assertFalse(any(
                    isinstance(cell.value, str) and cell.value and not cell.value.strip()
                    for ws in wb.worksheets for row in ws.iter_rows() for cell in row
                ))
                self.assertEqual(len(wb.active._images), 1)
            finally:
                wb.close()


class PdfExportTests(unittest.TestCase):
    def test_pdf_export_uses_xlsx_sibling_path(self):
        with tempfile.TemporaryDirectory() as directory:
            xlsx = Path(directory) / "document.xlsx"
            xlsx.write_bytes(b"xlsx")

            def fake_run(args, **kwargs):
                Path(kwargs["env"]["EASY_FULFILL_PDF_PATH"]).write_bytes(b"%PDF-1.4")
                return subprocess.CompletedProcess(args, 0, "", "")

            with patch("quotation_statement.subprocess.run", side_effect=fake_run):
                pdf = export_xlsx_to_pdf(xlsx)
            self.assertEqual(pdf, xlsx.with_suffix(".pdf"))
            self.assertTrue(pdf.is_file())
            self.assertGreater(pdf.stat().st_size, 0)
            self.assertIn("completed", (Path(directory) / "pdf_export.log").read_text(encoding="utf-8"))

    def test_pdf_export_keeps_generated_pdf_when_excel_cleanup_fails(self):
        with tempfile.TemporaryDirectory() as directory:
            xlsx = Path(directory) / "document.xlsx"
            xlsx.write_bytes(b"xlsx")

            def fake_run(args, **kwargs):
                Path(kwargs["env"]["EASY_FULFILL_PDF_PATH"]).write_bytes(b"%PDF-1.4")
                return subprocess.CompletedProcess(args, 1, "cleanup=Workbook.Close hresult=-2147418111", "")

            with patch("quotation_statement.subprocess.run", side_effect=fake_run):
                pdf = export_xlsx_to_pdf(xlsx)
            self.assertTrue(pdf.is_file())
            log = (Path(directory) / "pdf_export.log").read_text(encoding="utf-8")
            self.assertIn("completed_with_cleanup_warning", log)
            self.assertIn("-2147418111", log)


class NaverProductTests(unittest.TestCase):
    def test_sale_product_list_uses_discounted_price_and_smartstore_only(self):
        response = {
            "contents": [{
                "channelProducts": [
                    {
                        "channelServiceType": "STOREFARM",
                        "channelProductNo": 123,
                        "name": "STM32 개발 보드",
                        "salePrice": 12000,
                        "discountedPrice": 10000,
                    },
                    {
                        "channelServiceType": "WINDOW",
                        "channelProductNo": 456,
                        "name": "쇼핑윈도 상품",
                        "salePrice": 9000,
                    },
                ]
            }],
            "totalElements": 1,
        }
        with patch.object(naver_commerce, "_post_json", return_value=response):
            products = naver_commerce.fetch_sale_products("token")
        self.assertEqual(products, [{
            "product_no": "123",
            "name": "STM32 개발 보드",
            "sale_price": 12000,
            "discounted_price": 10000,
            "price": 10000,
        }])


class NaverTransactionStatementOrderTests(unittest.TestCase):
    def _detail(self, *, status="DELIVERED", total=9_999, quantity=2):
        return {
            "order": {
                "orderId": "202608120001",
                "ordererName": "홍길동",
                "paymentDate": "2026-08-12T10:20:30.000+09:00",
            },
            "productOrder": {
                "productOrderId": "202608120002",
                "productOrderStatus": status,
                "productName": "STM32 개발보드",
                "productOption": "색상: 검정",
                "quantity": quantity,
                "totalPaymentAmount": total,
                "shippingAddress": {"name": "홍수취"},
            },
        }

    def test_build_transaction_statement_order_prefers_orderer_and_keeps_exact_paid_amount(self):
        order = naver_commerce.build_transaction_statement_order(
            "202608120001", [self._detail()]
        )
        self.assertEqual(order["customer_name"], "홍길동")
        self.assertEqual(order["payment_date"], "2026-08-12T10:20:30.000+09:00")
        self.assertEqual(order["shipping_gross"], 0)
        self.assertEqual(order["items"], [{
            "product_order_id": "202608120002",
            "name": "STM32 개발보드",
            "specification": "색상: 검정",
            "quantity": 2,
            "gross_unit_price": 5_000,
            "gross_amount": 9_999,
            "status": "DELIVERED",
        }])

    def test_build_transaction_statement_order_falls_back_to_receiver_name(self):
        detail = self._detail()
        detail["order"]["ordererName"] = ""
        order = naver_commerce.build_transaction_statement_order("202608120001", [detail])
        self.assertEqual(order["customer_name"], "홍수취")

    def test_build_transaction_statement_order_blocks_cancelled_or_returned_order(self):
        with self.assertRaisesRegex(ValueError, "취소·반품·교환 완료"):
            naver_commerce.build_transaction_statement_order(
                "202608120001", [self._detail(status="RETURNED")]
            )

    def test_fetch_transaction_statement_order_resolves_product_order_ids_then_details(self):
        detail = self._detail()
        with patch.object(naver_commerce, "fetch_product_order_ids_of_order", return_value=["202608120002"]) as ids:
            with patch.object(naver_commerce, "fetch_product_order_details", return_value=[detail]) as details:
                order = naver_commerce.fetch_order_for_transaction_statement("token", "202608120001")
        ids.assert_called_once_with("token", "202608120001")
        details.assert_called_once_with("token", ["202608120002"])
        self.assertEqual(order["items"][0]["gross_amount"], 9_999)

    def test_build_transaction_statement_order_includes_shipping_once_per_package(self):
        first = self._detail()
        first["productOrder"].update({
            "packageNumber": "묶음배송-1",
            "deliveryFeeAmount": 3_000,
            "sectionDeliveryFee": 200,
            "deliveryDiscountAmount": 500,
        })
        second = self._detail(total=20_000, quantity=1)
        second["productOrder"].update({
            "productOrderId": "202608120003",
            "packageNumber": "묶음배송-1",
            "deliveryFeeAmount": 3_000,
            "sectionDeliveryFee": 200,
            "deliveryDiscountAmount": 500,
        })
        order = naver_commerce.build_transaction_statement_order("202608120001", [first, second])
        self.assertEqual(order["shipping_gross"], 2_700)


class CoupangTransactionStatementOrderTests(unittest.TestCase):
    @staticmethod
    def _money(amount):
        return {"currencyCode": "KRW", "units": amount, "nanos": 0}

    def _sheet(self, *, shipment_box_id="853138000000000001", order_price=14_000,
               discount=500, shipping=2_500, remote=0):
        return {
            "shipmentBoxId": shipment_box_id,
            "orderId": "500000596",
            "paidAt": "2026-08-12T10:20:30+09:00",
            "status": "FINAL_DELIVERY",
            "orderer": {"name": "홍길동"},
            "receiver": {"name": "수취인"},
            "shippingPrice": self._money(shipping),
            "remotePrice": self._money(remote),
            "orderItems": [{
                "sellerProductName": "개발 보드",
                "sellerProductItemName": "검정 / 1개",
                "vendorItemName": "개발 보드, 검정",
                "shippingCount": 2,
                "orderPrice": self._money(order_price),
                "discountPrice": self._money(discount),
                "cancelCount": 0,
                "holdCountForCancel": 0,
                "canceled": False,
            }],
        }

    def test_build_transaction_statement_order_uses_actual_discounted_goods_and_shipping(self):
        order = coupang_commerce.build_transaction_statement_order("500000596", [self._sheet()])
        self.assertEqual(order["customer_name"], "홍길동")
        self.assertEqual(order["payment_date"], "2026-08-12T10:20:30+09:00")
        self.assertEqual(order["shipping_gross"], 2_500)
        self.assertEqual(order["items"][0]["gross_amount"], 13_500)
        result = calculate_document(
            [ItemInput("개발 보드", gross_unit_price=6_750, quantity=2, gross_amount_override=13_500)],
            order_import=True, shipping_gross_override=2_500,
        )
        self.assertEqual(result.grand_total, 16_000)

    def test_build_transaction_statement_order_deduplicates_shipping_per_box(self):
        first = self._sheet()
        second = self._sheet()
        second["orderItems"][0]["sellerProductName"] = "추가 상품"
        order = coupang_commerce.build_transaction_statement_order("500000596", [first, second])
        self.assertEqual(order["shipping_gross"], 2_500)

    def test_build_transaction_statement_order_includes_remote_shipping_per_box(self):
        order = coupang_commerce.build_transaction_statement_order(
            "500000596", [self._sheet(shipping=2_500, remote=3_000)]
        )
        self.assertEqual(order["shipping_gross"], 5_500)

    def test_build_transaction_statement_order_blocks_partial_cancellation(self):
        sheet = self._sheet()
        sheet["orderItems"][0]["holdCountForCancel"] = 1
        with self.assertRaisesRegex(ValueError, "부분취소"):
            coupang_commerce.build_transaction_statement_order("500000596", [sheet])

    def test_build_transaction_statement_order_rejects_ambiguous_currency(self):
        sheet = self._sheet()
        sheet["orderItems"][0]["discountPrice"]["currencyCode"] = "USD"
        with self.assertRaisesRegex(ValueError, "KRW"):
            coupang_commerce.build_transaction_statement_order("500000596", [sheet])

    def test_fetch_transaction_statement_order_uses_single_order_api(self):
        sheet = self._sheet()
        with patch.object(coupang_commerce, "_get", return_value={"data": [sheet]}) as get:
            order = coupang_commerce.fetch_order_for_transaction_statement(
                "A00000001", "access", "secret", "500000596"
            )
        get.assert_called_once_with(
            "/v2/providers/openapi/apis/api/v5/vendors/A00000001/500000596/ordersheets",
            {}, "access", "secret",
        )
        self.assertEqual(order["items"][0]["gross_amount"], 13_500)


class WorkbookGenerationTests(unittest.TestCase):
    def test_manual_discount_is_used_for_generated_document(self):
        with tempfile.TemporaryDirectory() as directory:
            path, result = generate_document(
                "거래명세서",
                "테스트 소속",
                "홍길동",
                date(2026, 8, 11),
                [item(30_000, 4)],
                discount_rate_override=Decimal("0.07"),
                templates_dir=TEMPLATES,
                output_dir=directory,
            )
            wb = load_workbook(path, data_only=False)
            try:
                self.assertEqual(result.discount_rate, Decimal("0.07"))
                self.assertEqual(result.grand_total, 111_600)
                self.assertEqual(wb.active["H7"].value, 111_600)
            finally:
                wb.close()

    def test_quote_amount_label_preserves_the_original_font_sizes(self):
        with tempfile.TemporaryDirectory() as directory:
            path, _ = generate_document(
                "견적서", "테스트 소속", "홍길동", date(2026, 8, 11), [item(30_000)],
                templates_dir=TEMPLATES, output_dir=directory,
            )
            wb = load_workbook(path, rich_text=True)
            try:
                self.assertEqual(path.name, "견적서_테스트_소속_홍길동_2026.08.11.xlsx")
                self.assertEqual(wb.active["B26"].value, "기본 납기 : 결제 후 즉시")
                label = wb.active["B12"].value
                self.assertEqual(str(label), "견적금액 (공급가액 + 세액) ")
                self.assertEqual([part.font.sz for part in label], [14.0, 9.0])
            finally:
                wb.close()

    def test_statement_total_label_preserves_the_original_font_sizes(self):
        with tempfile.TemporaryDirectory() as directory:
            path, _ = generate_document(
                "거래명세서", "테스트 소속", "홍길동", date(2026, 8, 11), [item(30_000)],
                templates_dir=TEMPLATES, output_dir=directory,
            )
            wb = load_workbook(path, rich_text=True)
            try:
                self.assertEqual(path.name, "거래명세서_테스트_소속_홍길동_2026.08.11.xlsx")
                label = wb.active["B7"].value
                self.assertEqual(str(label), "합계 (부가세 포함) ")
                self.assertEqual([part.font.sz for part in label], [14.0, 9.0])
            finally:
                wb.close()

    def test_statement_allows_blank_organization_and_keeps_recipient_text_clean(self):
        with tempfile.TemporaryDirectory() as directory:
            path, _ = generate_document(
                "거래명세서", "", "홍길동", date(2026, 8, 11), [item(30_000)],
                templates_dir=TEMPLATES, output_dir=directory,
            )
            wb = load_workbook(path, data_only=False)
            try:
                self.assertEqual(path.name, "거래명세서_미입력_홍길동_2026.08.11.xlsx")
                self.assertEqual(wb.active["B5"].value, "홍길동님 귀하")
            finally:
                wb.close()

    def test_statement_writes_imported_actual_shipping_fee_as_a_line(self):
        with tempfile.TemporaryDirectory() as directory:
            path, result = generate_document(
                "거래명세서", "테스트 소속", "홍길동", date(2026, 8, 11),
                [ItemInput("주문 상품", gross_unit_price=5_000, quantity=2, gross_amount_override=9_999)],
                order_import=True,
                shipping_gross_override=2_500,
                templates_dir=TEMPLATES,
                output_dir=directory,
            )
            wb = load_workbook(path, data_only=False)
            try:
                ws = wb.active
                self.assertEqual(ws["B11"].value, "배송비")
                self.assertEqual(ws["M11"].value, 2_273)
                self.assertEqual(ws["T11"].value, 227)
                self.assertEqual(result.grand_total, 12_499)
            finally:
                wb.close()

    def test_output_filename_adds_a_number_when_the_base_name_exists(self):
        with tempfile.TemporaryDirectory() as directory:
            args = ("견적서", "테스트 소속", "홍길동", date(2026, 8, 11), [item(30_000)])
            first, _ = generate_document(*args, templates_dir=TEMPLATES, output_dir=directory)
            second, _ = generate_document(*args, templates_dir=TEMPLATES, output_dir=directory)
            self.assertEqual(first.name, "견적서_테스트_소속_홍길동_2026.08.11.xlsx")
            self.assertEqual(second.name, "견적서_테스트_소속_홍길동_2026.08.11_(1).xlsx")

    def test_shipping_uses_an_added_line_when_base_rows_are_full(self):
        items = [item(10_000, name=f"소액 품목 {n}") for n in range(1, 8)]
        with tempfile.TemporaryDirectory() as directory:
            path, result = generate_document(
                "견적서", "테스트 소속", "홍길동", date(2026, 8, 11), items,
                templates_dir=TEMPLATES, output_dir=directory,
            )
            wb = load_workbook(path, data_only=False)
            try:
                ws = wb.active
                self.assertEqual(ws["B23"].value, "배송비")
                self.assertEqual(ws["N23"].value, 2_727)
                self.assertEqual(ws["V23"].value, 273)
                self.assertEqual(ws["R24"].value, result.supply_total)
                self.assertEqual(ws["B28"].value, "결제 방법 : 직접 거래")
            finally:
                wb.close()

    def test_quote_adds_eighth_item_row_and_keeps_totals(self):
        items = [item(30_000, name=f"긴 상품명 {n} 테스트") for n in range(1, 9)]
        with tempfile.TemporaryDirectory() as directory:
            path, result = generate_document(
                "견적서", "테스트 소속", "홍길동", date(2026, 8, 11), items,
                payment_method="네이버", delivery_term="발주 후 3일", templates_dir=TEMPLATES, output_dir=directory,
            )
            wb = load_workbook(path, data_only=False)
            try:
                ws = wb.active
                self.assertEqual(ws["B23"].value, "긴 상품명 8 테스트")
                self.assertEqual(ws["R24"].value, result.supply_total)
                self.assertEqual(ws["V24"].value, result.tax_total)
                self.assertEqual(ws["R12"].value, result.grand_total)
                self.assertEqual(ws["R12"].number_format, WON_NUMBER_FORMAT)
                self.assertEqual(ws["R24"].number_format, WON_NUMBER_FORMAT)
                self.assertEqual(ws["V24"].number_format, WON_NUMBER_FORMAT)
                self.assertEqual(result.discount_rate, Decimal("0.05"))
                self.assertEqual(ws["B28"].value, "결제 방법 : 네이버 스토어 거래")
                self.assertEqual(ws["B27"].value, "기본 납기 : 발주 후 3일")
                self.assertIn("B23:F23", {str(r) for r in ws.merged_cells.ranges})
                self.assertTrue(ws.row_dimensions[29].hidden)
                self.assertNotIn("B29:V29", {str(r) for r in ws.merged_cells.ranges})
                self.assertEqual(str(ws.print_area), "'스토어 견적서'!$A$1:$V$28")
                self.assertEqual(ws.page_setup.fitToWidth, 1)
                self.assertEqual(ws.page_setup.fitToHeight, 0)
                self.assertEqual(len(ws._images), 1)
            finally:
                wb.close()

    def test_statement_adds_seventh_item_row_and_keeps_totals(self):
        items = [item(40_000, name=f"품목 {n}") for n in range(1, 8)]
        with tempfile.TemporaryDirectory() as directory:
            path, result = generate_document(
                "거래명세서", "테스트 소속", "홍길동", date(2026, 8, 11), items,
                templates_dir=TEMPLATES, output_dir=directory,
            )
            wb = load_workbook(path, data_only=False)
            try:
                ws = wb.active
                self.assertEqual(ws["B16"].value, "품목 7")
                self.assertEqual(ws["Q17"].value, result.supply_total)
                self.assertEqual(ws["T17"].value, result.tax_total)
                self.assertEqual(ws["H7"].value, result.grand_total)
                self.assertEqual(ws["B6"].value, "2026.08.11")
                self.assertIn("B16:F16", {str(r) for r in ws.merged_cells.ranges})
                self.assertEqual(str(ws.print_area), "'거래명세서'!$A$1:$W$19")
                self.assertEqual(ws.page_setup.fitToHeight, 0)
                self.assertEqual(len(ws._images), 1)
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
